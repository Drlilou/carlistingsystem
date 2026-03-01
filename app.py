from flask import Flask, render_template, request, redirect, url_for, flash, session,jsonify,Response,current_app
from models import db, Car, CarImage  # Import CarImage model

from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
import csv
from io import StringIO
from models import db, Car, CarImage 
import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font,PatternFill
from io import BytesIO
from datetime import datetime
import os

app = Flask(__name__)

app.secret_key = os.environ.get("SECRET_KEY", "your-very-secret-key-here")  # Change in production!

# 🔑 Simple admin credentials (in real app: use DB)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD_HASH = generate_password_hash("password123")  # Default password: password123

# 🔑 Simple admin credentials (in real app: use DB)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD_HASH = generate_password_hash("password123")  # Default password: password123


app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///cars.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "connect_args": {"check_same_thread": False}
}
db.init_app(app)


# Configure upload folder
UPLOAD_FOLDER = 'static/images'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.before_request
def create_tables():
    db.create_all()

@app.route('/')
def index():
    page = request.args.get('page', 1, type=int)

    query = Car.query

    # SEARCH
    if request.args.get('name'):
        query = query.filter(Car.name.ilike(f"%{request.args['name']}%"))
    if request.args.get('fuel'):
        query = query.filter(Car.fuel.ilike(f"%{request.args['fuel']}%"))
    if request.args.get('type'):
        query = query.filter(Car.car_type.ilike(f"%{request.args['type']}%"))
    if request.args.get('seats'):
        query = query.filter(Car.seats == request.args.get('seats'))

    pagination = query.paginate(page=page, per_page=9, error_out=False)

    return render_template(
        'index.html',
        cars=pagination.items,
        pagination=pagination
    )

# ===== LOGIN =====
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session['admin_logged_in'] = True
            flash('Logged in successfully!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid credentials. Try again.', 'error')
    return render_template('auth/login.html')
# ===== ADD CAR =====





@app.route('/admin/add-car', methods=['GET', 'POST'])
def add_car():
    if not session.get('admin_logged_in'):
        flash('Please log in first.', 'warning')
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
            # === Basic car data ===
            name = request.form['name'].strip()
            price = int(request.form['price'])
            seats = int(request.form['seats'])
            fuel = request.form['fuel'].strip()
            car_type = request.form['type'].strip()
            description = request.form.get('description', '').strip()

            if not all([name, price > 0, seats > 0, fuel, car_type]):
                flash('All car fields are required.', 'error')
                return render_template('admin/add_car.html')

            # === Handle images ===
            image_sources = []

            # 1. Process uploaded files
            image_files = request.files.getlist('image_files')
            for img in image_files:
                if img and img.filename != '' and allowed_file(img.filename):
                    filename = secure_filename(img.filename)
                    base, ext = os.path.splitext(filename)
                    counter = 1
                    while os.path.exists(os.path.join(UPLOAD_FOLDER, filename)):
                        filename = f"{base}_{counter}{ext}"
                        counter += 1
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    img.save(filepath)
                    image_sources.append(f"images/{filename}")

            # 2. Process URLs from textarea
            image_urls_raw = request.form.get('image_urls', '').strip()
            if image_urls_raw:
                urls = [url.strip() for url in image_urls_raw.split('\n') if url.strip()]
                image_sources.extend(urls)

            # 3. Validate at least one image
            if not image_sources:
                flash('At least one image (upload or URL) is required.', 'error')
                return render_template('admin/add_car.html')

            # === Save to database ===
            new_car = Car(
                name=name,
                price_per_day=price,
                seats=seats,
                fuel=fuel,
                car_type=car_type,
                description=description
            )
            db.session.add(new_car)
            db.session.flush()  # Get new_car.id

            # Add all images
            for img_url in image_sources:
                db.session.add(CarImage(image_url=img_url, car_id=new_car.id))

            db.session.commit()
            flash(f'✅ Car "{name}" added with {len(image_sources)} image(s)!', 'success')
            return redirect(url_for('admin_dashboard'))

        except ValueError:
            flash('Invalid number format.', 'error')
        except Exception as e:
            db.session.rollback()
            flash('An error occurred. Please try again.', 'error')
            print(f"[ERROR] Adding car: {e}")

    return render_template('admin/add_car.html')

# ===== ADMIN DASHBOARD =====
@app.route('/')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        flash('Please log in first.', 'warning')
        return redirect(url_for('login'))
    return render_template('admin/dashboard.html')

@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))



@app.route('/api/car/<int:car_id>')
def get_car_detail(car_id):
    car = Car.query.get_or_404(car_id)
    return jsonify({
        "name": car.name,
        "price": car.price_per_day,
        "seats": car.seats,
        "fuel": car.fuel,
        "type": car.car_type,
        "description": car.description,
        "images": [img.image_url for img in car.images]  # Include image URLs
    })

@app.route('/upload', methods=['POST'])
def upload_image():
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    file = request.files['file']
    
    if file:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Assuming you provide car_id via form data to associate the image
        car_id = request.form.get('car_id')
        new_image = CarImage(image_url=os.path.join('uploads', filename), car_id=car_id)
        db.session.add(new_image)
        db.session.commit()
        
        return jsonify({"message": "Image uploaded successfully", "image_url": new_image.image_url}), 201


@app.template_filter('asset_url')
def asset_url(image_path):
    """
    Converts a stored image path into a proper URL:
    - If it starts with http(s) → return as-is (external)
    - Otherwise → treat as local file under /static/
    """
    if image_path.startswith(('http://', 'https://')):
        return image_path
    else:
        return url_for('static', filename=image_path)


# ===== EDIT CAR =====

@app.route('/admin/car/<int:car_id>/edit', methods=['GET', 'POST'])
def edit_car(car_id):
    if not session.get('admin_logged_in'):
        flash('Admin access required.', 'error')
        return redirect(url_for('login'))

    car = Car.query.get_or_404(car_id)

    if request.method == 'POST':
        try:
            # Update basic fields
            car.name = request.form['name'].strip()
            car.price_per_day = int(request.form['price'])
            car.seats = int(request.form['seats'])
            car.fuel = request.form['fuel'].strip()
            car.car_type = request.form['type'].strip()
            car.description = request.form.get('description', '').strip()

            if not all([car.name, car.price_per_day > 0, car.seats > 0, car.fuel, car.car_type]):
                flash('All car fields are required.', 'error')
                return render_template('admin/edit_car.html', car=car)

            # Decide whether to replace or keep existing images
            replace_images = 'replace_images' in request.form

            new_image_sources = []

            # 1. Handle uploaded files
            image_files = request.files.getlist('image_files')
            for img in image_files:
                if img and img.filename != '' and allowed_file(img.filename):
                    filename = secure_filename(img.filename)
                    base, ext = os.path.splitext(filename)
                    counter = 1
                    while os.path.exists(os.path.join(UPLOAD_FOLDER, filename)):
                        filename = f"{base}_{counter}{ext}"
                        counter += 1
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    img.save(filepath)
                    new_image_sources.append(f"images/{filename}")

            # 2. Handle image URLs
            image_urls_raw = request.form.get('image_urls', '').strip()
            if image_urls_raw:
                urls = [url.strip() for url in image_urls_raw.split('\n') if url.strip()]
                new_image_sources.extend(urls)

            # 3. Save images
            if replace_images:
                # Delete old images
                CarImage.query.filter_by(car_id=car.id).delete()
                # Add all new ones
                for url in new_image_sources:
                    db.session.add(CarImage(image_url=url, car_id=car.id))
            else:
                # Only add new ones (keep old)
                if new_image_sources:
                    for url in new_image_sources:
                        db.session.add(CarImage(image_url=url, car_id=car.id))
                # If no new images, keep existing

            # Ensure at least one image exists
            total_images = CarImage.query.filter_by(car_id=car.id).count()
            if total_images == 0 and not new_image_sources:
                flash('At least one image is required.', 'error')
                return render_template('admin/edit_car.html', car=car)

            db.session.commit()
            flash(f'✅ Car "{car.name}" updated successfully!', 'success')
            return redirect(url_for('admin_dashboard'))

        except ValueError:
            flash('Invalid number format.', 'error')
        except Exception as e:
            db.session.rollback()
            flash('Error updating car. Please try again.', 'error')
            print(f"[ERROR] Editing car: {e}")

    # GET request: render form
    return render_template('admin/edit_car.html', car=car)


# ===== DELETE CAR =====
@app.route('/admin/car/<int:car_id>/delete', methods=['POST'])
def delete_car(car_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 403

    car = Car.query.get_or_404(car_id)
    car_name = car.name

    try:
        # Delete associated images first (optional, but clean)
        CarImage.query.filter_by(car_id=car_id).delete()
        db.session.delete(car)
        db.session.commit()
        return jsonify({'success': True, 'message': f'Car "{car_name}" deleted'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete car'}), 500
        





@app.route('/export/cars')
def export_cars():

    # =========================
    # FILTER LOGIC
    # =========================
    query = Car.query

    if request.args.get('name'):
        query = query.filter(Car.name.ilike(f"%{request.args['name']}%"))
    if request.args.get('fuel'):
        query = query.filter(Car.fuel.ilike(f"%{request.args['fuel']}%"))
    if request.args.get('type'):
        query = query.filter(Car.car_type.ilike(f"%{request.args['type']}%"))
    if request.args.get('seats'):
        try:
            query = query.filter(Car.seats == int(request.args.get('seats')))
        except (ValueError, TypeError):
            pass

    cars = query.all()

    # =========================
    # CREATE EXCEL FILE
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Cars Export"

    # =========================
    # ADD LOGO (FIXED PATH)
    # =========================
    logo_path = os.path.join(
        current_app.root_path,
        "static",
        "images",
        "src",
        "logo.png"
    )

    if os.path.exists(logo_path):
        logo = ExcelImage(logo_path)
        logo.width = 140
        logo.height = 70
        ws.add_image(logo, "A1")
    else:
        print("Logo not found:", logo_path)


    # =========================
    # ADD DATE
    # =========================
    export_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["G2"] = f"Exported on: {export_date}"
    ws["G2"].font = Font(bold=True)

    # =========================
    # TABLE HEADER
    # =========================
    start_row = 5
    headers = [
        "ID", "Name", "Price / Day", "Seats",
        "Fuel", "Car Type", "Description"#, "Image URLs"
    ]

    header_fill = PatternFill(
        start_color="FF0000",  # Red
        end_color="FF0000",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"  # White text
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    """for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)"""

    # =========================
    # TABLE DATA
    # =========================
    row = start_row + 1
    for car in cars:
        image_urls = "; ".join(img.image_url for img in car.images)

        ws.append([
            car.id,
            car.name,
            car.price_per_day,
            car.seats,
            car.fuel,
            car.car_type,
            car.description or "",
            #image_urls
        ])
        row += 1

    # =========================
    # AUTO COLUMN WIDTH
    # =========================
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    # =========================
    # SEND FILE
    # =========================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"cars_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

if __name__ == "__main__":

    app.run(debug=True)
